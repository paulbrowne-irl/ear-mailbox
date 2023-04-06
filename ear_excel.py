import logging
import os.path
import shutil
from pandas.core.frame import DataFrame


import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook

import settings


'''
Clear the  output file, so we can reuse the formatting
'''
def clear_excel_output_file():

    print ("\nCLEARING OLD EXCEL EMAIL REPORT")

    #Make a backup of the original file
    counter =1
    while(os.path.exists(str(counter)+settings.EMAIL_REPORT_FILE)):
        logging.debug("Backup file "+str(counter)+settings.EMAIL_REPORT_FILE+" exists, increment and try again")
        counter +=1

    # copy over the template
    shutil.copyfile(settings.EXCEL_TEMPLATE,settings.EMAIL_REPORT_FILE)
    logging.debug("Created new backup file from template")

    #Open Sheet using Python
    workbook = load_workbook(filename=settings.EMAIL_REPORT_FILE)
    sheet = workbook.active

    #Now delete everything until we are only left with the header row
    # continuously delete row 2 until there
    # is only a single row left over 
    # that contains column names 
    while(sheet.max_row > 1):
        # this method removes the row 2
        logging.debug("deleting row")
        sheet.delete_rows(2)

        #update to user
        print(".", end ='')

    #Save the result
    workbook.save(filename=settings.EMAIL_REPORT_FILE)
    workbook.close