import logging
import os.path
import shutil
from pandas.core.frame import DataFrame

import win32com.client
import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook

import settings
    
'''
EAR - Email AI Reader
Support for managing an outlook inbox - this could be a personal, or share one

Capture information from Outlook into an Excel file
'''

import logging
import os.path
import shutil
from pandas.core.frame import DataFrame


import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook

import settings


'''
Clear the  output file, so we can reuse the formatting
'''
def _do_analysis():

    print ("\nCLEARING OLD EXCEL EMAIL REPORT")

    #Make a backup of the original file
    counter =1
    while(os.path.exists(str(counter)+settings.EMAIL_DATA_DUMP)):
        logging.debug("Backup file "+str(counter)+settings.EMAIL_DATA_DUMP+" exists, increment and try again")
        counter +=1

    # copy over the template
    shutil.copyfile(settings.EXCEL_TEMPLATE,settings.EMAIL_DATA_DUMP)
    logging.debug("Created new backup file from template")

    #Open Sheet using Python
    workbook = load_workbook(filename=settings.EMAIL_DATA_DUMP)
    sheet = workbook.active


    #Save the result
    workbook.save(filename=settings.EMAIL_DATA_DUMP)
    workbook.close

# simple code to run from command line
if __name__ == '__main__':

    _do_analysis()
